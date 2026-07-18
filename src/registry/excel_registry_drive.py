"""
Registra facturas en Excel local y Drive con hipervínculos funcionales.

Local: usa file:/// (funciona en Excel de Windows)
Drive: usa https://drive.google.com/open?id=... (funciona en navegador)
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime


COLOR_HEADER      = "1F4E79"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_FILA_PAR    = "D6E4F0"
COLOR_LINK        = "1155CC"

def _borde_fino():
    lado = Side(style="thin", color="B0B0B0")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _estilo_header(ws):
    fill   = PatternFill("solid", fgColor=COLOR_HEADER)
    fuente = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = fuente
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _borde_fino()

def _estilo_fila(ws, fila_idx: int):
    es_par = (fila_idx % 2 == 0)
    fill   = PatternFill("solid", fgColor=COLOR_FILA_PAR) if es_par else None
    for cell in ws[fila_idx]:
        if fill:
            cell.fill = fill
        cell.border    = _borde_fino()
        cell.alignment = Alignment(vertical="center")

def _ajustar_anchos(ws, anchos: dict):
    for col_letra, ancho in anchos.items():
        ws.column_dimensions[col_letra].width = ancho

def _anchos_registro():
    return {
        "A": 12,   # Fecha
        "B": 28,   # Consorcio
        "C": 30,   # Proveedor
        "D": 15,   # RUC proveedor
        "E": 10,   # Total
        "F": 12,   # Abrir PDF
        "G": 14,   # Abrir Carpeta
    }

def _anchos_pendientes():
    return {
        "A": 20,   # Fecha procesamiento
        "B": 40,   # Nombre archivo
        "C": 45,   # Motivo
    }


def _abrir_o_crear_registro(ruta_excel: Path) -> tuple:
    COLUMNAS = ["Fecha", "Consorcio", "Proveedor", "RUC proveedor", "Total",
                "Abrir PDF", "Abrir Carpeta"]

    ruta_excel.parent.mkdir(parents=True, exist_ok=True)

    if ruta_excel.exists():
        wb = load_workbook(ruta_excel)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturas"
        ws.append(COLUMNAS)
        ws.row_dimensions[1].height = 18
        _estilo_header(ws)
        _ajustar_anchos(ws, _anchos_registro())

    return wb, ws


def _abrir_o_crear_pendientes(ruta_excel: Path) -> tuple:
    COLUMNAS = ["Fecha procesamiento", "Nombre archivo", "Motivo"]

    ruta_excel.parent.mkdir(parents=True, exist_ok=True)

    if ruta_excel.exists():
        wb = load_workbook(ruta_excel)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pendientes"
        ws.append(COLUMNAS)
        ws.row_dimensions[1].height = 18
        _estilo_header(ws)
        _ajustar_anchos(ws, _anchos_pendientes())

    return wb, ws


def _parsear_fecha_ordenamiento(fecha_str: str) -> datetime:
    """Convierte 'DD/MM/YYYY' a datetime para poder ordenar. Fechas inválidas
    o vacías van al final."""
    try:
        return datetime.strptime(str(fecha_str).strip(), "%d/%m/%Y")
    except (ValueError, TypeError):
        return datetime.max


def _ordenar_registro_por_fecha(ws):
    """
    Reordena todas las filas de datos (desde la fila 2) por la columna
    Fecha (A), de más antigua a más reciente, preservando los hipervínculos
    de las columnas F (PDF) y G (Carpeta).
    """
    max_fila = ws.max_row
    if max_fila < 3:
        return  # 0 o 1 fila de datos: nada que ordenar

    filas = []
    for fila_idx in range(2, max_fila + 1):
        valores = [ws.cell(fila_idx, col).value for col in range(1, 6)]  # A-E

        celda_pdf = ws.cell(fila_idx, 6)
        celda_carpeta = ws.cell(fila_idx, 7)

        filas.append({
            "valores": valores,
            "pdf_texto": celda_pdf.value,
            "pdf_link": celda_pdf.hyperlink.target if celda_pdf.hyperlink else None,
            "carpeta_texto": celda_carpeta.value,
            "carpeta_link": celda_carpeta.hyperlink.target if celda_carpeta.hyperlink else None,
        })

    filas.sort(key=lambda f: _parsear_fecha_ordenamiento(f["valores"][0]))
    #filas.sort(key=lambda f: _parsear_fecha_ordenamiento(f["valores"][0]), reverse=True)
    #cambiar orden de fecha excel registro para que quede de mas reciente a mas antiguo, para que se vea primero lo mas reciente


    # Borrar filas de datos actuales y reescribir en el orden ya ordenado
    ws.delete_rows(2, max_fila - 1)

    for i, f in enumerate(filas):
        fila_idx = i + 2
        for col, valor in enumerate(f["valores"], start=1):
            ws.cell(fila_idx, col, valor)

        celda_pdf = ws.cell(fila_idx, 6, f["pdf_texto"])
        if f["pdf_link"]:
            celda_pdf.hyperlink = f["pdf_link"]
            celda_pdf.font = Font(color=COLOR_LINK, underline="single")
        else:
            celda_pdf.alignment = Alignment(horizontal="center")

        celda_carpeta = ws.cell(fila_idx, 7, f["carpeta_texto"])
        if f["carpeta_link"]:
            celda_carpeta.hyperlink = f["carpeta_link"]
            celda_carpeta.font = Font(color=COLOR_LINK, underline="single")
        else:
            celda_carpeta.alignment = Alignment(horizontal="center")

        _estilo_fila(ws, fila_idx)
        ws.row_dimensions[fila_idx].height = 15


def registrar_factura(
    ruta_excel: str,
    fila: dict,
    nombre_consorcio: str = "",
    ruta_pdf_final: str = "",
    ruta_carpeta: str = "",
    url_pdf_drive: str = "",
    url_carpeta_drive: str = "",
    es_drive: bool = False,
):
    """
    Agrega una fila al Excel con hipervínculos.

    Args:
        ruta_excel: ruta al .xlsx
        fila: dict con Fecha, Consorcio, Proveedor, RUC proveedor, Total
        ruta_pdf_final: ruta local al PDF (para file:///)
        ruta_carpeta: ruta local a la carpeta (para file:///)
        url_pdf_drive: URL de Drive del PDF (https://drive.google.com/open?id=...)
        url_carpeta_drive: URL de Drive de la carpeta
        es_drive: si True, usar URLs de Drive en lugar de file:///
    """
    ruta_excel = Path(ruta_excel)
    wb, ws = _abrir_o_crear_registro(ruta_excel)

    fila_idx = ws.max_row + 1

    # Columnas A–E: datos
    fecha_str = fila.get("Fecha", "")
    if fecha_str and "-" in str(fecha_str):
        try:
            partes = str(fecha_str).split("-")
            fecha_str = f"{partes[2]}/{partes[1]}/{partes[0]}"
        except:
            pass

    ws.cell(fila_idx, 1, fecha_str)
    ws.cell(fila_idx, 2, fila.get("Consorcio", ""))
    ws.cell(fila_idx, 3, fila.get("Proveedor", ""))
    ws.cell(fila_idx, 4, fila.get("RUC proveedor", ""))
    ws.cell(fila_idx, 5, fila.get("Total", ""))

    # Columna F: Abrir PDF
    celda_pdf = ws.cell(fila_idx, 6, "Abrir")
    if es_drive and url_pdf_drive:
        celda_pdf.value = "📄 Abrir"
        celda_pdf.hyperlink = url_pdf_drive
        celda_pdf.font = Font(color=COLOR_LINK, underline="single")
    elif not es_drive and ruta_pdf_final:
        celda_pdf.value = "📄 Abrir"
        uri = Path(ruta_pdf_final).as_uri()
        celda_pdf.hyperlink = uri
        celda_pdf.font = Font(color=COLOR_LINK, underline="single")
    else:
        celda_pdf.value = "Abrir"
        celda_pdf.alignment = Alignment(horizontal="center")

    # Columna G: Abrir Carpeta
    celda_carpeta = ws.cell(fila_idx, 7, "Carpeta")
    if es_drive and url_carpeta_drive:
        celda_carpeta.value = "📁 Carpeta"
        celda_carpeta.hyperlink = url_carpeta_drive
        celda_carpeta.font = Font(color=COLOR_LINK, underline="single")
    elif not es_drive and ruta_carpeta:
        celda_carpeta.value = "📁 Carpeta"
        uri = Path(ruta_carpeta).as_uri()
        celda_carpeta.hyperlink = uri
        celda_carpeta.font = Font(color=COLOR_LINK, underline="single")
    else:
        celda_carpeta.value = "Carpeta"
        celda_carpeta.alignment = Alignment(horizontal="center")

    _estilo_fila(ws, fila_idx)
    ws.row_dimensions[fila_idx].height = 15

    _ordenar_registro_por_fecha(ws)

    wb.save(ruta_excel)


def registrar_pendiente(
    ruta_excel_pendientes: str,
    nombre_archivo: str,
    motivo: str = "",
):
    ruta_excel = Path(ruta_excel_pendientes)
    wb, ws = _abrir_o_crear_pendientes(ruta_excel)

    fila_idx = ws.max_row + 1
    fecha_hoy = datetime.now().strftime("%d/%m/%Y %H:%M")

    ws.cell(fila_idx, 1, fecha_hoy)
    ws.cell(fila_idx, 2, nombre_archivo)
    ws.cell(fila_idx, 3, motivo)

    _estilo_fila(ws, fila_idx)
    ws.row_dimensions[fila_idx].height = 15

    wb.save(ruta_excel)