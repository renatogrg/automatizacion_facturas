"""
Registra cada factura procesada en el Excel del consorcio correspondiente,
y las no resueltas en Registro_Pendientes.xlsx

Formato de columnas (local y Drive):
  Fecha | Consorcio | Proveedor | RUC proveedor | Total | Abrir PDF | Abrir Carpeta

Los hipervínculos "Abrir PDF" y "Abrir Carpeta" usan rutas absolutas con file:///
Solo funcionan como hipervínculo en el Excel LOCAL (Drive no admite file:///).
En el Excel de Drive las celdas muestran el texto igual pero sin hipervínculo activo,
lo cual es el comportamiento esperado ya que Drive usa URLs propias.
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime


# ─── Estilos ────────────────────────────────────────────────────────────────

COLOR_HEADER      = "1F4E79"   # azul oscuro
COLOR_HEADER_FONT = "FFFFFF"   # blanco
COLOR_FILA_PAR    = "D6E4F0"   # azul muy claro
COLOR_LINK        = "1155CC"   # azul hipervínculo estándar

def _borde_fino():
    lado = Side(style="thin", color="B0B0B0")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def _estilo_header(ws):
    fill   = PatternFill("solid", fgColor=COLOR_HEADER)
    fuente = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = fuente
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _borde_fino()

def _estilo_fila(ws, fila_idx: int):
    """Aplica fondo alternado y borde a cada celda de la fila."""
    es_par = (fila_idx % 2 == 0)
    fill   = PatternFill("solid", fgColor=COLOR_FILA_PAR) if es_par else None
    for cell in ws[fila_idx]:
        if fill:
            cell.fill = fill
        cell.border    = _borde_fino()
        cell.alignment = Alignment(vertical="center")

def _ajustar_anchos(ws, anchos: dict):
    for col_letra, ancho in anchos.items():
        ws.column_dimensions[col_letra].width = ancho

def _anchos_registro():
    return {
        "A": 12,   # Fecha
        "B": 28,   # Consorcio
        "C": 30,   # Proveedor
        "D": 15,   # RUC proveedor
        "E": 10,   # Total
        "F": 12,   # Abrir PDF
        "G": 14,   # Abrir Carpeta
    }

def _anchos_pendientes():
    return {
        "A": 20,   # Fecha procesamiento
        "B": 40,   # Nombre archivo
        "C": 45,   # Motivo
    }


# ─── Crear / abrir workbook ──────────────────────────────────────────────────

def _abrir_o_crear_registro(ruta_excel: Path) -> tuple:
    """Abre el Excel si existe, o crea uno nuevo con cabecera y estilos."""
    COLUMNAS = ["Fecha", "Consorcio", "Proveedor", "RUC proveedor", "Total",
                "Abrir PDF", "Abrir Carpeta"]

    ruta_excel.parent.mkdir(parents=True, exist_ok=True)

    if ruta_excel.exists():
        wb = load_workbook(ruta_excel)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Facturas"
        ws.append(COLUMNAS)
        ws.row_dimensions[1].height = 18
        _estilo_header(ws)
        _ajustar_anchos(ws, _anchos_registro())

    return wb, ws


def _abrir_o_crear_pendientes(ruta_excel: Path) -> tuple:
    COLUMNAS = ["Fecha procesamiento", "Nombre archivo", "Motivo"]

    ruta_excel.parent.mkdir(parents=True, exist_ok=True)

    if ruta_excel.exists():
        wb = load_workbook(ruta_excel)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Pendientes"
        ws.append(COLUMNAS)
        ws.row_dimensions[1].height = 18
        _estilo_header(ws)
        _ajustar_anchos(ws, _anchos_pendientes())

    return wb, ws


# ─── Funciones públicas ───────────────────────────────────────────────────────

def registrar_factura(
    ruta_excel: str,
    fila: dict,
    nombre_consorcio: str = "",
    ruta_pdf_final: str = "",
    ruta_carpeta: str = "",
    es_drive: bool = False,
):
    """
    Agrega una fila al Excel de registro del consorcio.

    Args:
        ruta_excel:       ruta al archivo .xlsx (local o Drive)
        fila:             dict con Fecha, Consorcio, Proveedor, RUC proveedor, Total
        nombre_consorcio: nombre del consorcio (informativo)
        ruta_pdf_final:   ruta absoluta al PDF ya organizado (para hipervínculo)
        ruta_carpeta:     ruta a la carpeta donde quedó el PDF (para hipervínculo)
        es_drive:         True → no generar hipervínculos file:/// (no funcionan en Drive)
    """
    ruta_excel = Path(ruta_excel)
    wb, ws = _abrir_o_crear_registro(ruta_excel)

    fila_idx = ws.max_row + 1

    # Columnas A–E: datos
    fecha_str = fila.get("Fecha", "")
    if fecha_str and "-" in str(fecha_str):
        try:
            partes = str(fecha_str).split("-")
            fecha_str = f"{partes[2]}/{partes[1]}/{partes[0]}"
        except Exception:
            pass

    ws.cell(fila_idx, 1, fecha_str)
    ws.cell(fila_idx, 2, fila.get("Consorcio", ""))
    ws.cell(fila_idx, 3, fila.get("Proveedor", ""))
    ws.cell(fila_idx, 4, fila.get("RUC proveedor", ""))
    ws.cell(fila_idx, 5, fila.get("Total", ""))

    # Columna F: Abrir PDF
    celda_pdf = ws.cell(fila_idx, 6, "Abrir PDF")  # Sin emoji
    if ruta_pdf_final and not es_drive:
        uri = Path(ruta_pdf_final).as_uri()          # file:///C:/FACTURAS/...
        celda_pdf.hyperlink = uri
        celda_pdf.font      = Font(color=COLOR_LINK, underline="single")
    else:
        celda_pdf.alignment = Alignment(horizontal="center")

    # Columna G: Abrir Carpeta
    celda_carpeta = ws.cell(fila_idx, 7, "Abrir Carpeta")  # Sin emoji
    if ruta_carpeta and not es_drive:
        uri = Path(ruta_carpeta).as_uri()
        celda_carpeta.hyperlink = uri
        celda_carpeta.font      = Font(color=COLOR_LINK, underline="single")
    else:
        celda_carpeta.alignment = Alignment(horizontal="center")

    _estilo_fila(ws, fila_idx)
    ws.row_dimensions[fila_idx].height = 15

    wb.save(ruta_excel)


def registrar_pendiente(
    ruta_excel_pendientes: str,
    nombre_archivo: str,
    motivo: str = "",
):
    """
    Registra una factura que no pudo clasificarse.
    Se llama tanto para el Excel local como para el de Drive.
    """
    ruta_excel = Path(ruta_excel_pendientes)
    wb, ws = _abrir_o_crear_pendientes(ruta_excel)

    fila_idx = ws.max_row + 1
    fecha_hoy = datetime.now().strftime("%d/%m/%Y %H:%M")

    ws.cell(fila_idx, 1, fecha_hoy)
    ws.cell(fila_idx, 2, nombre_archivo)
    ws.cell(fila_idx, 3, motivo)

    _estilo_fila(ws, fila_idx)
    ws.row_dimensions[fila_idx].height = 15

    wb.save(ruta_excel)